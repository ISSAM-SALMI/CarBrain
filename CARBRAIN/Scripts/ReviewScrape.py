import undetected_chromedriver as uc
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
import os
import csv
from path import PATH_PARK_XLSX, PATH_PARK_CSV


class CarScraper:
    def __init__(self, headless=True):
        self.options = uc.ChromeOptions()
        if headless:
            self.options.add_argument('--headless')
        self.options.add_argument('--disable-gpu')
        self.options.add_argument('--no-sandbox')
        self.options.add_argument('--disable-dev-shm-usage')
        self.driver = uc.Chrome(options=self.options)

    def open_page(self, url):
        try:
            self.driver.get(url)
        except TimeoutException as e:
            print(f"Error: Unable to load the page {url} - {e}")

    def open_page_with_numberkey(self, page_number, key):
        url = f"https://www.parkers.co.uk/car-reviews/{key}/?page={page_number}"
        self.open_page(url)

    def get_info_of_page(self):
        try:
            WebDriverWait(self.driver, 30).until(
                EC.presence_of_element_located((By.CLASS_NAME, 'model-panel'))
            )
            return self.driver.find_elements(By.CLASS_NAME, 'model-panel')
        except TimeoutException:
            return []

    def verify_xlsx_file(self, path=PATH_PARK_XLSX):
        if os.path.exists(path):
            try:
                wb = load_workbook(path)
                ws = wb.active
            except Exception as e:
                print(f"Erreur lors de l'ouverture du fichier : {e}")
                print("Création d'un nouveau fichier Excel...")
                os.remove(path)
                wb = Workbook()
                ws = wb.active
                ws.append(["Car", "Quote", "Ranking", "NewPrice", "Pros+", "Cons-", 'link'])
                wb.save(path)
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Car", "Quote", "Ranking", "NewPrice", "Pros+", "Cons-", "link"])
            wb.save(path)

    def verify_csv_file(self, path=PATH_PARK_CSV):
        """Verify if the CSV file exists and create it with headers if necessary."""
        if not os.path.exists(path):
            with open(path, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(["Car", "Quote", "Ranking", "NewPrice", "Pros+", "Cons-", 'link'])

    def extract_info_of_single_page(self, page_number=1, key="small-city"):
        self.open_page_with_numberkey(page_number, key)
        articles = self.get_info_of_page()
        if not articles:
            return {}

        try:
            return {
                "Car": [
                    article.find_element(By.CLASS_NAME, "panel__primary-link").text
                    for article in articles
                ],
                "Quote": [
                    article.find_element(By.CLASS_NAME, "model-panel__strapline").text
                    for article in articles
                ],
                "Ranking": [
                    article.find_element(By.CLASS_NAME, "star-rating__stars__value").text
                    for article in articles
                ],
                "NewPrice": [
                    article.find_element(By.CLASS_NAME, "model-panel__price__value").text
                    for article in articles
                ],
                "Pros+": [
                    "\n".join(
                        li_item.text for li_item in 
                        article.find_elements(By.CLASS_NAME, "info-box__item")[0].find_elements(By.TAG_NAME, "li")
                    ) if len(article.find_elements(By.CLASS_NAME, "info-box__item")) > 0 else "N/A"
                    for article in articles
                ],
                "Cons-": [
                    "\n".join(
                        li_item.text for li_item in 
                        article.find_elements(By.CLASS_NAME, "info-box__item")[1].find_elements(By.TAG_NAME, "li")
                    ) if len(article.find_elements(By.CLASS_NAME, "info-box__item")) > 1 else "N/A"
                    for article in articles
                ],
            }
        except NoSuchElementException as e:
            print(f"Error while extracting data: {e}")
            return {}

    def save_car_details(self, car_url, page_number=1, key="small-city", path_xlsx=PATH_PARK_XLSX, path_csv=PATH_PARK_CSV):
        self.verify_xlsx_file()
        self.verify_csv_file()

        data = self.extract_info_of_single_page(page_number, key)
        if not data:
            print(f"No data extracted for {car_url}")
            return

        carsinfo = list(data.values()) + [car_url]

        # Sauvegarde dans Excel
        wb = load_workbook(path_xlsx)
        ws = wb.active
        for row in zip(*data.values()):
            ws.append(list(row) + [car_url])
        wb.save(path_xlsx)

        # Sauvegarde dans CSV
        with open(path_csv, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            for row in zip(*data.values()):
                writer.writerow(list(row) + [car_url])

    def __del__(self):
        self.driver.quit()


if __name__ == "__main__":
    cars = [
        "small-city", "hatchback", "saloon", "estate", "4x4",
        "mpv-people-carrier", "suv", "coupe", "convertible",
        "family", "fast-sports", "electric-hybrid",
    ]
    scraper = CarScraper()

    for car in cars:
        page_number = 1
        while True:
            try:
                data = scraper.save_car_details(page_number, car)
                if not data:
                    break
                print(f"Scraped page {page_number} for category {car}: {data}")
                page_number += 1
            except Exception as e:
                print(f"Stopping scraping for {car} at page {page_number} (Reason: {e})")
                break